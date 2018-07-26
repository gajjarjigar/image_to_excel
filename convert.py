from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter


backgroundColor = (0,) * 3

if __name__ == '__main__':
	pixelSize = 10
	image = Image.open('my_image.png')
	image = image.resize((int(image.size[0] / pixelSize), int(image.size[1] / pixelSize)), Image.NEAREST)
	rgb_im = image.convert('RGB')
	wb = openpyxl.Workbook()
	ws = wb.active
	m = 1
	for i in range(image.size[0]):
		n = 1
		for j in range(image.size[1]):
			r, g, b = rgb_im.getpixel((i, j))
			hex_color = 'FF%02x%02x%02x' % (r, g, b)
			hex_color = hex_color.upper()
			color_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
			ws.cell(row=n, column=m).value = ''
			ws.cell(row=n, column=m).fill = color_fill
			n += 1
		m += 1

	column_widths = [3] * image.size[0]
	for i, column_width in enumerate(column_widths):
		ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	wb.save('art.xlsx')
